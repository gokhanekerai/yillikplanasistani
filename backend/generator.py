import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from holidays import get_holidays_by_year
from week_generator import generate_school_calendar

def create_excel_plan(json_data: list, academic_year: str, custom_title: str, output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Yıllık Plan"

    # Sütun Genişliklerini Ayarla (A4 Yatay için optimize)
    col_widths = {
        'A': 5,   # Sıra
        'B': 10,  # Ay
        'C': 15,  # Hafta/Tarih
        'D': 6,   # Saat
        'E': 45,  # Kazanımlar
        'F': 25,  # Konular
        'G': 15,  # Yöntem
        'H': 15,  # Materyal
        'I': 15   # Açıklama
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # Başlık
    title_text = custom_title if custom_title else f"{academic_year} EĞİTİM ÖĞRETİM YILI YILLIK PLANI"
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = title_text
    title_cell.font = Font(name="Times New Roman", size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Başlık satır yüksekliği dinamik hesaplama (yaklaşık)
    lines = title_text.count('\n') + 1
    ws.row_dimensions[1].height = 25 * lines

    # Header Row
    headers = ["SIRA", "AY", "HAFTA / TARİH", "SAAT", "KAZANIMLAR", "KONULAR", "YÖNTEM/TEKNİK", "MATERYALLER", "AÇIKLAMA"]
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(name="Times New Roman", size=12, bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    # Verileri Yazma
    holidays_db = get_holidays_by_year(academic_year)
    weeks = generate_school_calendar(holidays_db)
    
    current_row = 4
    data_idx = 0

    base_font = Font(name="Times New Roman", size=12)
    base_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    holiday_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    holiday_font = Font(name="Times New Roman", size=12, bold=True, color="FFFFFF")
    
    partial_fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
    partial_font = Font(name="Times New Roman", size=12, bold=True, color="333333")

    for i, week in enumerate(weeks):
        # Hafta tarih metni ve ay
        date_text = f"{week['start'].day} - {week['end'].day} {week['end'].strftime('%B').upper()}"
        if week['start'].month != week['end'].month:
            date_text = f"{week['start'].day} {week['start'].strftime('%B').upper()} - {week['end'].day} {week['end'].strftime('%B').upper()}"
        
        # Aylar ingilizce çıkar, manuel türkçeye çevir
        months_tr = {
            "JANUARY": "OCAK", "FEBRUARY": "ŞUBAT", "MARCH": "MART", "APRIL": "NİSAN",
            "MAY": "MAYIS", "JUNE": "HAZİRAN", "JULY": "TEMMUZ", "AUGUST": "AĞUSTOS",
            "SEPTEMBER": "EYLÜL", "OCTOBER": "EKİM", "NOVEMBER": "KASIM", "DECEMBER": "ARALIK"
        }
        for eng, tr in months_tr.items():
            date_text = date_text.replace(eng, tr)
        
        ay_text = months_tr[week['end'].strftime('%B').upper()]

        # Eğer hafta tamamen tatilse
        if not week['hasNormalClass']:
            h_names = " / ".join(list(set([d['holidayName'] for d in week['days'] if d['isHoliday']])))
            
            # A, B, C, D sütunlarını doldur
            ws.cell(row=current_row, column=1, value="").border = thin_border
            ws.cell(row=current_row, column=2, value=ay_text).border = thin_border
            ws.cell(row=current_row, column=2).alignment = center_align
            ws.cell(row=current_row, column=2).font = base_font
            
            ws.cell(row=current_row, column=3, value=date_text).border = thin_border
            ws.cell(row=current_row, column=3).alignment = center_align
            ws.cell(row=current_row, column=3).font = base_font
            
            ws.cell(row=current_row, column=4, value="TATİL").border = thin_border
            ws.cell(row=current_row, column=4).alignment = center_align
            ws.cell(row=current_row, column=4).font = base_font

            # Geri kalanını birleştir ve kırmızıya boya
            cell = ws.cell(row=current_row, column=5, value=h_names)
            ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=9)
            
            for col in range(1, 10):
                c = ws.cell(row=current_row, column=col)
                c.border = thin_border
                c.fill = holiday_fill
                c.font = holiday_font
            
            current_row += 1
            continue

        # Normal Hafta (JSON verisini kullan)
        row_data = {}
        if data_idx < len(json_data):
            row_data = json_data[data_idx]
            data_idx += 1
            
        kazanimlar = row_data.get("kazanimlar", "")
        konular = row_data.get("konular", "")
        yontem = row_data.get("yontem", "")
        materyaller = row_data.get("materyaller", "")
        aciklama = row_data.get("aciklama", "")

        # Kısmi tatil var mı?
        h_names = list(set([d['holidayName'] for d in week['days'] if d['isHoliday']]))
        if h_names:
            extra = "🎉 " + " / ".join(h_names)
            aciklama = aciklama + "\n" + extra if aciklama else extra

        # Sütun değerlerini ata
        ws.cell(row=current_row, column=1, value=i+1)
        ws.cell(row=current_row, column=2, value=ay_text)
        ws.cell(row=current_row, column=3, value=date_text)
        ws.cell(row=current_row, column=4, value=5) # Default saat
        ws.cell(row=current_row, column=5, value=kazanimlar)
        ws.cell(row=current_row, column=6, value=konular)
        ws.cell(row=current_row, column=7, value=yontem)
        ws.cell(row=current_row, column=8, value=materyaller)
        
        # Açıklama hücresi
        aciklama_cell = ws.cell(row=current_row, column=9, value=aciklama)
        if h_names:
            aciklama_cell.fill = partial_fill
            aciklama_cell.font = partial_font

        # Stil ve Border
        for col in range(1, 10):
            c = ws.cell(row=current_row, column=col)
            c.border = thin_border
            if col in [1, 2, 3, 4]:
                c.alignment = center_align
            else:
                c.alignment = base_align
            if not (col == 9 and h_names):
                c.font = base_font

        current_row += 1

    wb.save(output_path)
    return output_path

"""
Gerçekçi bir yıllık plan Excel dosyası oluşturur (test amaçlı).
A sütununda hafta aralıkları, diğer sütunlarda ders konuları.
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Yıllık Plan"

# Başlık satırı
headers = ["HAFTA", "KONU", "KAZANIMLAR", "AÇIKLAMA"]
header_fill = PatternFill(start_color="FF2C3E50", end_color="FF2C3E50", fill_type="solid")
header_font = Font(bold=True, color="FFFFFFFF", size=11)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

ws.row_dimensions[1].height = 25

# Haftalık veriler — bazıları tatil dönemine denk geliyor
weeks_data = [
    ("14-18 EYLÜL",       "Türkçe: Okuma Becerileri",          "5.1.1, 5.1.2", ""),
    ("21-25 EYLÜL",       "Matematik: Doğal Sayılar",           "5.1.1",        ""),
    ("28 EYLÜL - 2 EKİM", "Fen: Hücre ve Organeller",          "5.1.1",        ""),
    ("5-9 EKİM",          "Sosyal: Ülkemizin Kaynakları",      "5.1.1",        ""),
    ("12-16 EKİM",        "İngilizce: Greetings",              "5.1.1",        ""),
    ("19-23 EKİM",        "Türkçe: Yazma Becerileri",          "5.2.1",        ""),
    ("26-30 EKİM",        "Matematik: Kesirler",                "5.2.1",        ""),  # 28-29 Cumhuriyet Bayramı
    ("2-6 KASIM",         "Fen: Vücudumuzdaki Sistemler",      "5.2.1",        ""),
    ("9-13 KASIM",        "Sosyal: Türkiye'nin İklimi",        "5.2.1",        ""),
    ("16-20 KASIM",       "ARA TATİL",                         "",             ""),  # 16-20 Kasım Ara Tatil
    ("23-27 KASIM",       "Matematik: Geometri",                "5.3.1",        ""),
    ("30 KASIM - 4 ARALIK","Türkçe: Okuma Anlama",             "5.3.1",        ""),
    ("7-11 ARALIK",       "Fen: Kuvvet ve Hareket",            "5.3.1",        ""),
    ("14-18 ARALIK",      "Sosyal: Demokrasi",                 "5.3.1",        ""),
    ("21-25 ARALIK",      "İngilizce: Daily Routines",         "5.3.1",        ""),
    ("28 ARALIK - 1 OCAK","Matematik: Veri ve Olasılık",       "5.4.1",        ""),  # 1 Ocak Yılbaşı
    ("4-8 OCAK",          "Türkçe: Şiir Analizi",              "5.4.1",        ""),
    ("11-15 OCAK",        "Fen: Madde ve Değişim",             "5.4.1",        ""),
    ("18-22 OCAK",        "1. DÖNEM SONU SINAVI",              "",             ""),
    # Yarıyıl Tatili: 25 Ocak - 5 Şubat arası
    ("8-12 ŞUBAT",        "Matematik: Kesirler (Tekrar)",       "5.1.1",        ""),
    ("15-19 ŞUBAT",       "Türkçe: Metin Türleri",             "5.1.2",        ""),
    ("22-26 ŞUBAT",       "Fen: Işık ve Ses",                  "5.2.1",        ""),
    ("1-5 MART",          "Sosyal: İnsan Hakları",             "5.2.1",        ""),
    ("8-12 MART",         "2. DÖNEM ARA TATİL",                "",             ""),  # 8-12 Mart Ara Tatil (Ramazan)
    ("15-19 MART",        "İngilizce: Shopping",               "5.2.2",        ""),
    ("22-26 MART",        "Matematik: Üçgenler",               "5.3.1",        ""),
    ("29 MART - 2 NİSAN", "Türkçe: Roman Okuma",               "5.3.2",        ""),
    ("5-9 NİSAN",         "Fen: Ekosistem",                    "5.3.1",        ""),
    ("12-16 NİSAN",       "Sosyal: Kurtuluş Savaşı",          "5.3.1",        ""),
    ("19-23 NİSAN",       "Ulusal Egemenlik Haftası",          "",             ""),  # 23 Nisan
    ("26-30 NİSAN",       "Matematik: Olasılık",               "5.4.1",        ""),
    ("3-7 MAYIS",         "Türkçe: Yazılı Anlatım",            "5.4.1",        ""),
    ("10-14 MAYIS",       "Fen: Güneş Sistemi",                "5.4.1",        ""),
    ("17-21 MAYIS",       "KURBAN BAYRAMI TATİLİ",             "",             ""),  # 15-19 Mayıs + 19 Mayıs
    ("24-28 MAYIS",       "Sosyal: Coğrafya Tekrar",           "5.4.2",        ""),
    ("31 MAYIS - 4 HAZİRAN","Matematik: Genel Tekrar",         "5.4.2",        ""),
    ("7-11 HAZİRAN",      "Türkçe: Genel Tekrar",              "5.4.2",        ""),
    ("14-18 HAZİRAN",     "Fen: Genel Tekrar",                 "5.4.2",        ""),
    ("21-25 HAZİRAN",     "YILSONU SINAVI",                    "",             ""),
]

row_fill_odd  = PatternFill(start_color="FFF8FAFC", end_color="FFF8FAFC", fill_type="solid")
row_fill_even = PatternFill(start_color="FFEFF6FF", end_color="FFEFF6FF", fill_type="solid")

for i, (week, konu, kazanim, aciklama) in enumerate(weeks_data, 2):
    fill = row_fill_odd if i % 2 == 0 else row_fill_even
    data = [week, konu, kazanim, aciklama]
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=i, column=col, value=val)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[i].height = 20

# Sütun genişlikleri
ws.column_dimensions["A"].width = 28
ws.column_dimensions["B"].width = 45
ws.column_dimensions["C"].width = 20
ws.column_dimensions["D"].width = 20

output = "test_yillik_plan.xlsx"
wb.save(output)
print(f"OK: Test dosyasi olusturuldu: {output}")
